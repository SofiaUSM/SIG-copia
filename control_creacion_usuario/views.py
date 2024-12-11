from email.mime.image import MIMEImage
import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,update_session_auth_hash, login as auth_login, logout as auth_logout
from django.shortcuts import redirect
from django.contrib.auth.models import User
from django.http import HttpResponse,JsonResponse
from core.models import UserActivity
from formulario.models import ProtocoloSolicitud,Registro_designio
from django.utils import timezone
from openpyxl import Workbook
# Create your views here.
from django.views.decorators.csrf import csrf_exempt
from .forms import ImagenForm,PDFForm
from .models import Imagen_sig,PDF_sig
from django.shortcuts import get_object_or_404, redirect
from django.core.paginator import Paginator, Page
from datetime import datetime, timedelta
from django.utils.timezone import make_aware

from formulario.models import *

from django.db.models import Q,F, Sum,Count
import os
from django.contrib import messages
from django.utils.timezone import now
from arcgis.gis import GIS
from arcgis.features import FeatureLayer
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import pytz

gis = GIS("https://www.arcgis.com", "jimmi.gomez_munivalpo", "Jimgomez8718")

from .pdf_generator import *

def login(request):
    if request.user.is_authenticated:
        return redirect('control')

    if request.method == 'POST':
        email = request.POST.get('email')
        email = email.lower()
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)

        if user is not None:
            auth_login(request, user)
            return redirect('solicitude_llegadas')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'Login.html')

def download_excel(request):
    # Crear un nuevo libro de trabajo de Excel y agregar datos
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'PAGINA'
    sheet['B1'] = 'DEPARTAMENTO'
    sheet['C1'] = 'MES-DIA-AÑO-HORA'

    sheet.column_dimensions['A'].width = 30  # Ancho de la columna B
    sheet.column_dimensions['B'].width = 50  # Ancho de la columna C
    sheet.column_dimensions['C'].width = 30  # Ancho de la columna C


    row = 2  # Fila inicial para los datos
    activities = UserActivity.objects.all()

    for activity in activities:
        sheet.cell(row=row, column=1).value = activity.page
        sheet.cell(row=row, column=2).value = activity.departamento
        sheet.cell(row=row, column=3).value = activity.timestamp.strftime('%m-%d-%Y-%H:%M')
        row += 1  # Incrementar la fila para el próximo registro
    # Configurar la respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Historial_de_visitas.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)

    return response

def Historial_Visitas(request):
    Historial = UserActivity.objects.all()
    data = {
    'historial': Historial,
    }
    return render(request,'Historial_Visitas.html',data)

@login_required(login_url='/login/')
def solicitude_llegadas(request, dia_p=None):
    ESTADO = [
        ('RECIBIDO', 'RECIBIDO'),
        ('EN PROCESO', 'EN PROCESO'),
        ('EJECUTADO', 'EJECUTADO'),
        ('RECHAZADO', 'RECHAZADO')
    ]

    LIMITE_DE_DIA = {
        ('', ''),
        ('L', '1 - 2 Días Hábiles'),
        ('M', '2 - 4 Días Hábiles'),
        ('A', '3 - 5 Días Hábiles'),
        ('P', 'Especificar días manualmente'),
    }

    OPCIONES = {
        'ESTADO': ESTADO,
        'LIMITE_DE_DIA': LIMITE_DE_DIA
    }

    minutos = 0
    hora = 0
    usuarios = User.objects.all()

    # Filtrar las solicitudes según el usuario
    if request.user.is_superuser:
        solicitudes = ProtocoloSolicitud.objects.all()
    else:
        solicitudes = ProtocoloSolicitud.objects.filter(profesional=request.user)

    # Agregar información adicional sobre número de designios y días restantes
    solicitudes_data = []
    for solicitud in solicitudes:
        # Obtener el número de designios asociados a la solicitud
        numero_designios = Registro_designio.objects.filter(protocolo=solicitud).count()

        # Calcular los días restantes hasta la fecha límite
        if solicitud.fecha_T:
            dias_restantes = "Trabajo terminado"
        elif solicitud.fecha_L:
            # Calcular la diferencia en segundos
            total_segundos = (solicitud.fecha_L - now()).total_seconds()

            if total_segundos < 0:  # Si el tiempo ya pasó
                total_segundos = abs(total_segundos)
                dias_pasados = int(total_segundos // (24 * 3600))
                horas_pasadas = int((total_segundos % (24 * 3600)) // 3600)
                minutos_pasados = int((total_segundos % 3600) // 60)

                if dias_pasados > 0 or horas_pasadas > 0 or minutos_pasados > 0:
                    dias_restantes = f"Pasada por {dias_pasados} días, {horas_pasadas} horas y {minutos_pasados} minutos"
                else:
                    dias_restantes = "El tiempo límite ha pasado recientemente"
            else:  # Tiempo restante
                # Calcular días hábiles restantes
                dias_habiles_restantes = calcular_dias_habiles(now().date(), solicitud.fecha_L.date())
                horas_restantes = int((total_segundos % (24 * 3600)) // 3600)
                minutos_restantes = int((total_segundos % 3600) // 60)
                

                if dias_habiles_restantes > 1:
                    dias_restantes = f"Te quedan {dias_habiles_restantes-1} días hábiles"
                elif horas_restantes > 0:
                    dias_restantes = f"Te quedan {horas_restantes} horas y {minutos_restantes} minutos"
                else:
                    dias_restantes = f"Te quedan {minutos_restantes} minutos"
        else:
            dias_restantes = "Sin fecha límite"

        # Agregar los datos al listado de solicitudes
        solicitudes_data.append({
            'solicitud': solicitud,
            'numero_designios': numero_designios,
            'dias_restantes': dias_restantes
        })

    data = {
        'OPCIONES': OPCIONES,
        'Solicitudes': solicitudes_data,  # Lista con la información adicional
        'Usuarios': usuarios,  # Lista de usuarios
    }

    return render(request, 'solicitude_llegadas.html', data)

def calcular_dias_habiles(fecha_inicio, fecha_fin):
    dias_habiles = 0
    dia_actual = fecha_inicio

    while dia_actual <= fecha_fin:
        # Si el día actual no es sábado (5) ni domingo (6), lo contamos
        if dia_actual.weekday() < 5:  # 0 = Lunes, ..., 4 = Viernes
            dias_habiles += 1
        dia_actual += timedelta(days=1)
    
    return dias_habiles

@login_required(login_url='/login/')
def control(request):
    # Obtener todas las solicitudes
    total_solicitudes = ProtocoloSolicitud.objects.count()

    # Contar solicitudes por estado
    estados = ProtocoloSolicitud.objects.values("estado").annotate(
        total=Count("estado")
    )
    # Convertir a un diccionario para facilitar el acceso
    estado_counts = {estado["estado"]: estado["total"] for estado in estados}

    # Asegurarse de que todos los estados estén representados, incluso si son 0
    estados_posibles = ["RECIBIDO", "EN PROCESO", "EJECUTADO", "RECHAZADO"]
    for estado in estados_posibles:
        estado_counts.setdefault(estado, 0)

    # Contar solicitudes por profesional SIG
    solicitudes_per_profesional = (
        ProtocoloSolicitud.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Crear listas para etiquetas y datos del gráfico
    labels = [
        f"{item['profesional__first_name']} {item['profesional__last_name']}"
        for item in solicitudes_per_profesional
    ]
    data = [item["total"] for item in solicitudes_per_profesional]

    # Convertir las listas a JSON para utilizarlas en JavaScript
    labels_json = json.dumps(labels)
    data_json = json.dumps(data)

    # Mapeo de tipo_limite a días máximos
    tipo_limite_days = {
        "L": 2,  # LIVIANA
        "M": 4,  # MEDIA
        "A": 5,  # ALTO
    }

    # Contar solicitudes por tipo_limite
    counts_per_tipo = (
        ProtocoloSolicitud.objects.filter(tipo_limite__in=tipo_limite_days.keys())
        .values("tipo_limite")
        .annotate(total=Count("tipo_limite"))
    )

    # Calcular el promedio ponderado de carga de trabajo
    weighted_sum = sum(
        tipo_limite_days[item["tipo_limite"]] * item["total"]
        for item in counts_per_tipo
    )
    total_tipo = sum(item["total"] for item in counts_per_tipo)
    average_carga_trabajo = round(weighted_sum / total_tipo, 2) if total_tipo > 0 else 0

    # Preparar datos de tipo_limite para la plantilla
    tipo_limite_stats = [
        {
            "tipo_limite": "ALTO",
            "dias_maximos": 5,
            "total": next(
                (
                    item["total"]
                    for item in counts_per_tipo
                    if item["tipo_limite"] == "A"
                ),
                0,
            ),
        },
        {
            "tipo_limite": "MEDIA",
            "dias_maximos": 4,
            "total": next(
                (
                    item["total"]
                    for item in counts_per_tipo
                    if item["tipo_limite"] == "M"
                ),
                0,
            ),
        },
        {
            "tipo_limite": "LIVIANA",
            "dias_maximos": 2,
            "total": next(
                (
                    item["total"]
                    for item in counts_per_tipo
                    if item["tipo_limite"] == "L"
                ),
                0,
            ),
        },
    ]

    context = {
        "total_solicitudes": total_solicitudes,
        "en_proceso": estado_counts.get("EN PROCESO", 0),
        "ejecutado": estado_counts.get("EJECUTADO", 0),
        "rechazado": estado_counts.get("RECHAZADO", 0),
        "labels_json": labels_json,
        "data_json": data_json,
        "average_carga_trabajo": average_carga_trabajo,
        "tipo_limite_stats": tipo_limite_stats,
    }

    return render(request, "Control.html", context)

def cambiar_contraseña(request):
    if request.method == 'POST':
            password1 = request.POST['password1']
            password2 = request.POST['password2']

            if password1 == password2:
                user = User.objects.get(id=request.user.id)
                user.set_password(password1)
                user.save()

                # Mantener al usuario autenticado después del cambio de contraseña
                update_session_auth_hash(request, user)

                return redirect('control')
            else:
                messages.error(request, 'Las contraseñas no coinciden. Inténtalo de nuevo.')

    return render(request, 'cambiar_contraseña.html')

def logout(request):
    auth_logout(request)
    return redirect('control')

def Gestion_imagen(request):
    if request.method == 'POST':
        form = ImagenForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Gestion_imagen')  # Redirige a la misma vista después de guardar la imagen
    else:
        form = ImagenForm()

    # Obtener todas las imágenes
    imagenes = Imagen_sig.objects.all()
    paginator = Paginator(imagenes, 6)  # Muestra 6 imágenes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Gestion_imagen.html', {'form': form, 'imagenes': imagenes,'page_obj': page_obj})

@csrf_exempt
def actualizar_estado(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        nuevo_estado = request.POST.get('estado')
        ArchivoProtocolo = request.POST.get('files')
        solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)
        tdc = None

        if nuevo_estado =="EJECUTADO" and solicitud.profesional != None:
            solicitud.estado = nuevo_estado
            solicitud.fecha_T = timezone.now()
            solicitud.save()
            tdc,tpr = Calculor_de_trabajo()

        elif nuevo_estado =="RECHAZADO":
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)
            solicitud.estado = nuevo_estado
           
            solicitud.save()
        else:
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)
            solicitud.estado = nuevo_estado
            solicitud.save()

        if  tdc:
            try:
                    # Establecer conexión con ArcGIS (ajusta esta parte según tu configuración)
                    feature_layer = FeatureLayer("https://services8.arcgis.com/5BaAHTQ4nRVz57H5/arcgis/rest/services/total_carga_laboral/FeatureServer/0")

                    # Consulta para seleccionar la característica
                    query = feature_layer.query(where="id = 0")

                    # Modificación del atributo 'Visita'
                    for feature in query.features:
                        feature.attributes['cumpli'] = tdc
                        feature.attributes['ejec'] = tpr
                    
                    # Aplicar cambios a la capa de ArcGIS
                    feature_layer.edit_features(updates=query.features)

            except Exception as e:
                # Manejo de errores
                    print(f"Error al realizar la consulta o la actualización")

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido'})

@csrf_exempt
def actualizar_profesional(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        nuevo_profesional_id = request.POST.get('profesional')
        motivo = request.POST.get('motivo')
        solicitud = get_object_or_404(ProtocoloSolicitud, id=solicitud_id)


        if not nuevo_profesional_id or not solicitud_id:
            return JsonResponse({'success': False, 'message': 'Necesita'})


        if solicitud.profesional:
            
            Reg = Registro_designio(
                objetivos = motivo,
                protocolo = solicitud,
                profesional = solicitud.profesional,
                
                )
            Reg.save()
            
        nuevo_profesional = get_object_or_404(User, id=nuevo_profesional_id)
        solicitud.profesional = nuevo_profesional
        solicitud.fecha_D = timezone.now()
        solicitud.save()

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido'})

def calcular_fecha_limite(inicio, dias):
    fecha_limite = inicio
    dias_restantes = dias

    while dias_restantes > 0:
        fecha_limite += timedelta(days=1)
        if fecha_limite.weekday() < 5:  # Lunes=0, Domingo=6
            dias_restantes -= 1
    
    return fecha_limite

@csrf_exempt
def actualizar_limite(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        tipo_limite = request.POST.get('nuevoLimite')
        fecha_completa = request.POST.get('fecha')  # Fecha completa enviada desde el frontend

        try:
            # Validar si la solicitud existe
            solicitud = ProtocoloSolicitud.objects.get(id=solicitud_id)

            if tipo_limite == 'P':
                if not fecha_completa:
                    return JsonResponse({'success': False, 'message': 'Debe proporcionar una fecha válida para el tipo P'})
                try:
                    # Convertir la fecha completa ISO 8601 a datetime
                    fecha_limite = datetime.strptime(fecha_completa, "%Y-%m-%dT%H:%M:%S.%fZ")
                    # Asegurarse de que sea consciente de la zona horaria UTC y convertir a la zona horaria local
                    fecha_limite = make_aware(fecha_limite, timezone=pytz.UTC)
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'Formato de fecha inválido'})
            else:
                # Calcular los días límite según el tipo
                if tipo_limite == 'A':
                    dias_limite = 5
                elif tipo_limite == 'M':
                    dias_limite = 4
                elif tipo_limite == 'L':
                    dias_limite = 1
                else:
                    return JsonResponse({'success': False, 'message': 'Tipo de límite no reconocido'})

                # Calcular la fecha límite basada en días
                fecha_limite = calcular_fecha_limite(timezone.now(), dias_limite)

            # Actualizar la solicitud con el nuevo tipo de límite y fecha límite
            solicitud.tipo_limite = tipo_limite
            solicitud.fecha_L = fecha_limite
            solicitud.save()

            return JsonResponse({'success': True, 'message': 'Límite actualizado correctamente'})

        except ProtocoloSolicitud.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'La solicitud no existe'})

    return JsonResponse({'success': False, 'message': 'Método no permitido'})
def eliminar_imagen(request, imagen_id):
    imagen = get_object_or_404(Imagen_sig, pk=imagen_id)
    
    if request.method == 'POST':
        # Guarda la ruta del archivo de la imagen
        ruta_archivo = imagen.archivo_adjunto.path
        # Elimina la imagen de la base de datos
        imagen.delete()
        # Elimina el archivo de la imagen del sistema de archivos
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
        return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes después de eliminar la imagen
    
    return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes si la solicitud no es de tipo POST

def Gestion_pdf(request):
    if request.method == 'POST':
        form = PDFForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Gestion_pdf')  # Redirige a la misma vista después de guardar la imagen
    else:
        form = PDFForm()

    # Obtener todas las imágenes
    pdf = PDF_sig.objects.all()
    paginator = Paginator(pdf, 6)  # Muestra 6 imágenes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Gestion_PDF.html', {'form': form, 'PDF': pdf,'page_obj': page_obj})

def Calculor_de_trabajo():

    total_ejecutado = ProtocoloSolicitud.objects.filter(
    estado="EJECUTADO",
    fecha_D__isnull=False).count()

    print("Número total de solicitudes válidas:", total_ejecutado)


    total_ejecutado_a_tiempo = ProtocoloSolicitud.objects.filter(
        estado="EJECUTADO",
        fecha_D__isnull=False,
        fecha_T__isnull=False,                                        
        fecha_T__lte=F('fecha_L')  # fecha_T menor o igual que fecha_L
    ).count()

    print("Número total de solicitudes válidas a tiempo:", total_ejecutado_a_tiempo)


    solicitudes = ProtocoloSolicitud.objects.filter(
        estado="EJECUTADO",
        fecha_D__isnull=False,
        fecha_T__isnull=False
    )

    # Calcular el total de segundos
    total_dias = 0
    for s in solicitudes:
        diferencia = s.fecha_D  -  s.fecha_T # Restar los DateTimeField, obteniendo un timedelta
        total_dias += round(diferencia.total_seconds() / 86400, 3)  # Convertir segundos a días y limitar a 3 decimales

    total_dias = round(total_dias,3)
    print("Total de días entre fecha_D y fecha_T:", total_dias)


            
    if total_ejecutado > 0:
        tpr = total_ejecutado /  total_dias
        tpr = round(tpr,3)

    else:
        tpr = 0  # Evitar división por cero si no hay solicitudes ejecutadas

    print("Tiempo Promedio de Resolución (TPR) en Dias:", tpr)

    if total_ejecutado_a_tiempo > 0:
        tdc = (total_ejecutado /   total_ejecutado_a_tiempo )*100
        tdc = round(tdc,3)

    else:
        tdc = 0

    print("Tasa de cumplimiento:", tdc)

    return tdc,tpr

@csrf_exempt  # Solo usar esto si estás probando; para producción, configura CSRF correctamente
def Envio_de_correo(request):
    if request.method == 'POST':
        user = request.user
        emails = json.loads(request.POST.get('emails', '[]'))  # Convertir JSON de vuelta a lista
        if user.is_superuser:
            # Obtener los datos
            message = request.POST.get('message', '')
            ficha_id = request.POST.get('ficha_id')
            Protocolo = ProtocoloSolicitud.objects.get(id=ficha_id)
            Protocolo.enviado_correo = True
            profesional = Protocolo.profesional
            Protocolo.estado = 'EN PROCESO'
            Protocolo.save()

            # Generar PDF (opcional, según tu lógica)
            buffer = Generar_PDF(ficha_id)

            # Configuración del correo
            
            superusers = User.objects.filter(is_superuser=True, ).exclude(username=user.username).values_list('email', flat=True)
            superuser_emails = list(superusers)
            
            mi_coreo = f'{user.username}@munivalpo.cl'
            mi_coreo = mi_coreo.strip()

            asunto = f'Solicitud N°{Protocolo.codigo}  Asignada'
            mensaje = MIMEMultipart()
            mensaje['From'] = mi_coreo
            destinatarios = list(set([profesional.email] + emails + superuser_emails))
            mensaje['To'] = ', '.join(destinatarios)
            mensaje['Subject'] = asunto

            bcc_destinatarios = [mi_coreo]


            # Cargar la firma
            firma_path = os.path.join('media/assets/Firma', f'{user.username}.png')
            if os.path.exists(firma_path):
                with open(firma_path, 'rb') as firma_file:
                    firma_img = MIMEImage(firma_file.read())
                    firma_img.add_header('Content-ID', '<firma>')
                    mensaje.attach(firma_img)

            # Crear el contenido del correo con la firma
            html_content = f"""
            <html>
                <body>
                    <p>{message}</p>
                    <br>
                    <img src="cid:firma" alt="Firma" width="600" height="auto" />

                </body>
            </html>
            """
            mensaje.attach(MIMEText(html_content, 'html'))

            # Adjuntar PDF generado
            archivo_pdf = buffer.getvalue()
            pdf_adjunto = MIMEApplication(archivo_pdf)
            pdf_adjunto.add_header('Content-Disposition', 'attachment', filename='Ficha_de_protocolo.pdf')
            mensaje.attach(pdf_adjunto)

            # Adjuntar archivos enviados
            archivos = request.FILES.getlist('files')  # Obtener todos los archivos enviados
            for archivo in archivos:
                archivo_adjunto = MIMEApplication(archivo.read())
                archivo_adjunto.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=archivo.name
                )
                mensaje.attach(archivo_adjunto)

            # Configuración del servidor SMTP
                    # Configuración del servidor SMTP
            smtp_server = 'mail.munivalpo.cl'
            smtp_port = 587
            smtp_usuario = f'servervalpo\\{user.username}'

            contraseña = encotra_contraseña(user.username)

            smtp_contrasena = contraseña

            # Enviar el correo
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_usuario, smtp_contrasena)
            server.sendmail(
                mi_coreo,
                destinatarios + bcc_destinatarios,  # Incluir los destinatarios normales y BCC
                mensaje.as_string()
            )
            server.quit()

            return JsonResponse({'success': True})
        else:
            
            message = request.POST.get('message', '')
            ficha_id = request.POST.get('ficha_id')
            Protocolo = ProtocoloSolicitud.objects.get(id=ficha_id)
            solicitante = Protocolo.corre_solicitante
            profesional = Protocolo.profesional
            Protocolo.estado = 'EJECUTADO'
            Protocolo.fecha_T = timezone.now()
            Protocolo.enviado_correo_t = True

            Protocolo.save()

            # Generar PDF (opcional, según tu lógica)
            buffer = Generar_PDF(ficha_id)
            
            superusers = User.objects.filter(is_superuser=True).exclude(username="osvaldo.moya").values_list('email', flat=True)
            superuser_emails = list(superusers)



            # Configuración del correo
            mi_coreo = f'{user.username}@munivalpo.cl'
            mi_coreo = mi_coreo.strip()
            asunto = f'Solicitud N° {Protocolo.codigo} Atendida'
            mensaje = MIMEMultipart()
            mensaje['From'] = mi_coreo
            mensaje['To'] = ', '.join([solicitante] + emails + superuser_emails)
            destinatarios = [solicitante] + emails + superuser_emails  # Remueve el correo del usuario aquí
            mensaje['Subject'] = asunto
            # Agregar el correo invisible en el campo BCC
            bcc_destinatarios = [mi_coreo]

            # Cargar la firma
            firma_path = os.path.join('media/assets/Firma', f'{user.username}.png')
            if os.path.exists(firma_path):
                with open(firma_path, 'rb') as firma_file:
                    firma_img = MIMEImage(firma_file.read())
                    firma_img.add_header('Content-ID', '<firma>')
                    mensaje.attach(firma_img)

            # Crear el contenido del correo con la firma
            html_content = f"""
            <html>
                <body>
                    <p>{message}</p>
                    <br>
                    <img src="cid:firma" alt="Firma" width="600" height="auto" />
                </body>
            </html>
            """
            mensaje.attach(MIMEText(html_content, 'html'))

            # Adjuntar PDF generado
            archivo_pdf = buffer.getvalue()
            pdf_adjunto = MIMEApplication(archivo_pdf)
            pdf_adjunto.add_header('Content-Disposition', 'attachment', filename='Ficha_de_protocolo.pdf')
            mensaje.attach(pdf_adjunto)

            # Adjuntar archivos enviados
            archivos = request.FILES.getlist('files')  # Obtener todos los archivos enviados
            for archivo in archivos:
                archivo_adjunto = MIMEApplication(archivo.read())
                archivo_adjunto.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=archivo.name
                )
                mensaje.attach(archivo_adjunto)

            # Configuración del servidor SMTP
            smtp_server = 'mail.munivalpo.cl'
            smtp_port = 587
            smtp_usuario = f'servervalpo\\{user.username}'

            contraseña = encotra_contraseña(user.username)

            smtp_contrasena = contraseña

            # Enviar el correo
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_usuario, smtp_contrasena)
            server.sendmail(
                f'{user.username}@munivalpo.cl', 
                destinatarios + bcc_destinatarios,  # Incluir los destinatarios normales y BCC
                mensaje.as_string()
            )
            server.quit()

            return JsonResponse({'success': True})


                    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@csrf_exempt
def vista_previa_reaccinacion(request, id):
    # Obtener el protocolo solicitado
    protocoloS = ProtocoloSolicitud.objects.get(id=id)
    
    # Filtrar los registros relacionados con el protocolo
    reg = Registro_designio.objects.filter(protocolo=protocoloS.id)
    
    # Construir una lista con los datos de cada registro
    registros_data = [
        {
            "id": registro.id,
            "Motivo": registro.objetivos,
            "Fecha": registro.fecha.strftime('%Y-%m-%d'),
            "Profesional_N": registro.profesional.first_name,
            "Profesional_N_1": registro.profesional.last_name,

        }
        for registro in reg
    ]
    
    # Preparar la respuesta
    data = {
        "registros": registros_data,
    }
    
    return JsonResponse(data)

def delegar_admin(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        updates = data.get('updates', [])

        for update in updates:
            user_id = update.get('user_id')
            is_superuser = update.get('is_superuser', False)

            try:
                user = User.objects.get(id=user_id)
                user.is_superuser = is_superuser
                user.save()
            except User.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': f'Usuario con ID {user_id} no encontrado.'}, status=404)

        return JsonResponse({'status': 'success', 'message': 'Usuarios actualizados correctamente.'})

    usuarios = User.objects.filter(is_staff=False)
    data = {
        'Usuarios': usuarios,
    }
    return render(request, 'admin.html', data)

def encotra_contraseña(usuario):
    secrets_file_path = 'pass.txt'

    with open(secrets_file_path, 'r') as file:
        secrets = file.readlines()

    # Buscar un usuario específico
    for secret in secrets:
        saved_user, saved_password = secret.strip().split(':')
        if saved_user == usuario:
            return saved_password

    return None 

@csrf_exempt
def resert_limite(request):
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            id_solicitud = data.get('id')
            solicitud = ProtocoloSolicitud.objects.get(id=id_solicitud)
            solicitud.tipo_limite = ''
            solicitud.fecha_L = None
            solicitud.save()


            return JsonResponse({'message': 'Solicitud reseteada con éxito.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Método no permitido.'}, status=405)
